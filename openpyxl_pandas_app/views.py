from django.shortcuts import render
from django.views import View

import openpyxl
import openpyxl.chart
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

# --- トップページ --- #
def index(request):
    return render(request, 'openpyxl_pandas.html')

# --- 10章 データのソートをしよう --- #
class Chap10(View):
    dir = 'openpyxl_pandas_app/static/excel/'
    template = "chap10.html"

    def dispatch(self, request, *args, **kwargs):
        if request.path.endswith('excel_read/'):
            return self.excel_read(request, *args, **kwargs)
        elif request.path.endswith('excel_write/'):
            return self.excel_write(request, *args, **kwargs)
        elif request.path.endswith('excel_update/'):
            return self.excel_update(request, *args, **kwargs)
        elif request.path.endswith('excel_delete/'):
            return self.excel_delete(request, *args, **kwargs)
        else:
            return self.top_page(request, *args, **kwargs)
    
    # トップページ
    def top_page(self, request, *args, **kwargs):
        context = {}
        return render(request, self.template, context)
    
    # Excelファイルの読み込み
    def excel_read(self, request, *args, **kwargs):
        wb = openpyxl.load_workbook(f"{self.dir}Chapter10_1.xlsx")
        ws = wb.active
        data = ws.values
        df = pd.DataFrame(data)
        pre = f"{df}"

        # コンテキスト
        context = { 'pre': pre }
        return render(request, self.template, context)
    
    # Excelファイルへの書き込み
    def excel_write(self, request, *args, **kwargs):
        # データフレームを作成する
        df = pd.DataFrame(data = {
            'ID': [1, 2, 3],
            '名前': ['田中一郎', '山田次郎', '加藤三郎'],
            '年齢': [20, 30, 40]
        })

        wb = Workbook()
        ws = wb.active

        # データフレームからリストを取得してExcelファイルに追加
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(f"{self.dir}Chapter10_2.xlsx")

        # コンテキスト
        msg = "Excelファイルへの書き込みを行いました"
        context = { 'msg': msg }
        return render(request, self.template, context)
    
    # Excelの更新
    def excel_update(self, request, *args, **kwargs):
        wb = openpyxl.load_workbook(f"{self.dir}Chapter10_3.xlsx")

        # 必要なシートを選択
        sheet_name = "Sheet1"
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Worksheet {sheet_name} does not exist.")
        sheet = wb[sheet_name]

        # シートのデータを Pandas データフレームに変換
        data = sheet.values
        columns = next(data)  # ヘッダーを取得
        df = pd.DataFrame(data, columns=columns)

        # データフレームを操作
        if not all(col in df.columns for col in ['国語', '数学', '英語']):
            raise KeyError("必要な列が見つかりません。列名を確認してください。")

        df['合計'] = df['国語'] + df['数学'] + df['英語']
        df = df.sort_values('合計', ascending=False)

        # 更新後のデータを OpenPyXL ワークシートに書き戻す
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 保存
        wb.save(f"{self.dir}Chapter10_3.xlsx")
        wb.close()

        # コンテキスト
        msg = "Excelファイルの更新を行いました"
        context = { 'msg': msg }
        return render(request, self.template, context)

    # Excelデータの削除
    def excel_delete(self, request, *args, **kwargs):

        # データフレームを作成する
        data = {
            'ID': [1, 2, 3],
            '名前': ['田中一郎', '山田次郎', '加藤三郎'],
            '年齢': [20, 30, 40]
        }
        df = pd.DataFrame(data)

        # 「ID」列を削除する
        df = df.drop('ID', axis=1)

        # 条件に合致するデータをフィルタリングする
        df = df[df['年齢'] >= 30]

        # Excelファイルの作成
        wb = Workbook()
        ws = wb.active

        # データフレームをExcelファイルに書き込む
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Excelファイルを閉じる
        wb.save(f"{self.dir}Chapter10_4.xlsx")
        wb.close()

        # コンテキスト
        msg = "Excelのデータ削除を行いました"
        context = { 'msg': msg }
        return render(request, self.template, context)

# --- 11章 データのグループ化とデータフレームの連結をしよう --- #
class Chap11(View):
    dir = 'openpyxl_pandas_app/static/excel/'
    template = "chap11.html"

    def dispatch(self, request, *args, **kwargs):
        if request.path.endswith('excel_read/'):
            return self.excel_read(request, *args, **kwargs)
        elif request.path.endswith('excel_select/'):
            return self.excel_select(request, *args, **kwargs)
        elif request.path.endswith('expression/'):
            return self.expression(request, *args, **kwargs)
        elif request.path.endswith('connection/'):
            return self.connection(request, *args, **kwargs)
        else:
            return self.top_page(request, *args, **kwargs)
    
    # トップページ
    def top_page(self, request, *args, **kwargs):
        context = {}
        return render(request, self.template, context)
    
    def excel_read(self, request, *args, **kwargs):
        # Excelファイルを読み込む
        df = pd.read_excel(f"{self.dir}Chapter11_1.xlsx", sheet_name='Sheet1')

        # 年代別に「得点」列の平均値を計算する
        result = df.groupby('年代')['得点'].mean()

        # コンテキスト
        pre = f"{result}"
        context = { 'pre': pre }
        return render(request, self.template, context)

    # 条件に応じたデータの抽出
    def excel_select(self, request, *args, **kwargs):
        df = pd.read_excel(f"{self.dir}Chapter11_2.xlsx")
        df = df[df['年代'] >= 40]

        # コンテキスト
        pre = f"{df}"
        context = { 'pre': pre }
        return render(request, self.template, context)
    
    # 正規表現によるパターンマッチング
    def expression(self, request, *args, **kwargs):
        # Excelファイルを読み込む
        df = pd.read_excel(f"{self.dir}Chapter11_3.xlsx", sheet_name='Sheet1')
        pattern = r'^[0-9]{3}-[0-9]{4}-[0-9]{4}$'
        df_filtered = df[df['電話番号'].str.contains(pattern, na=False)]

        # コンテキスト
        pre = f"{df_filtered}"
        context = { 'pre': pre }
        return render(request, self.template, context)
    
    # 複数のExcelファイルの連結
    def connection(self, request, *args, **kwargs):
        file_list = [f"{self.dir}Chapter11_4.xlsx", f"{self.dir}Chapter11_5.xlsx"]

        # 空のデータフレームを作成
        df = pd.DataFrame()

        # ファイルを一つずつ処理する
        for file_name in file_list:
            temp_df = pd.read_excel(file_name, header=None)
            temp_df.columns = ['名前', '年代', '得点']
            df = pd.concat([df, temp_df], ignore_index=True)

        # コンテキスト
        pre = f"{df}"
        context = { 'pre': pre }
        return render(request, self.template, context)

# --- 12章 openpyxlとpandasで実際にExcelファイルを作成しよう --- #
class Chap12(View):
    dir = 'openpyxl_pandas_app/static/excel/'
    template = "chap12.html"

    def dispatch(self, request, *args, **kwargs):
        if request.path.endswith('work-1-2/'):
            return self.work_1_2(request, *args, **kwargs)
        elif request.path.endswith('work-3-4/'):
            return self.work_3_4(request, *args, **kwargs)
        else:
            return self.top_page(request, *args, **kwargs)
    
    # トップページ
    def top_page(self, request, *args, **kwargs):
        context = {}
        return render(request, self.template, context)
    
    # Work(1)(2)
    def work_1_2(self, request, *args, **kwargs):
        # データフレームを作成する
        df = pd.DataFrame(data={
            '商品名': ['人気商品A', '商品B', '人気商品C', '人気商品A', '商品B', '商品D'],
            '売上金額': [1000, 2000, 3000, 4000, 5000, 6000],
            '売上コスト': [800, 1500, 2400, 800, 1500, 2200],
        })

        # 商品ごとに売上金額と売上コストを合計
        grouped = df.groupby('商品名').agg({'売上金額': 'sum', '売上コスト': 'sum'})
        grouped['利益率'] = ((grouped['売上金額'] - grouped['売上コスト']) / grouped['売上金額']) * 100

        # Excelファイルを作成する
        writer = pd.ExcelWriter(f"{self.dir}売上管理表.xlsx")

        # DataFrameオブジェクトをExcelファイルに書き込む
        grouped.to_excel(writer, sheet_name='売上管理', index=True)
    
        # データフレームから「人気」とついている商品を抽出
        popular_items = df[df['商品名'].str.contains('人気')]

        # Excelファイルに抽出結果を追加
        popular_items.to_excel(writer, sheet_name='人気商品', index=False)

        # Excelファイルを閉じる
        writer.close()

        text = f" \
        {grouped} \
        \n----------\n \
        {popular_items}"

        # コンテキスト
        pre = text
        context = { 'pre': pre }
        return render(request, self.template, context)
    
    # Work(3)(4)
    def work_3_4(self, request, *args, **kwargs):
        # Excelファイル「売上管理表.xlsx」を読み込む
        wb = openpyxl.load_workbook(f"{self.dir}売上管理表.xlsx")
        ws = wb.active
        data2 = ws.values
        cols = next(data2)
        df = pd.DataFrame(data2, columns=cols)

        # 新規シート「売上高グラフ」に棒グラフオブジェクトを作成する
        sheet1 = wb.create_sheet('売上高グラフ')
        chart1 = BarChart()

        # グラフのタイトル、データの範囲、カテゴリーの範囲を設定し追加する
        chart1.title = '商品別売上高'
        labels = Reference(ws, min_col=1, min_row=2, max_row=df.shape[0]+1)
        data = Reference(ws, min_col=2, min_row=2, max_row=df.shape[0]+1)

        # グラフにデータを追加する
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(labels)

        sheet1.add_chart(chart1, 'A1')

        # 利益率のグラフを作成して新しいシートに追加する
        sheet2 = wb.create_sheet('利益率グラフ')
        chart2 = BarChart()

        chart2.title = '商品別利益率'
        labels = Reference(ws, min_col=1, min_row=2, max_row=df.shape[0] + 1)
        data = Reference(ws, min_col=4, min_row=2, max_row=df.shape[0] + 1)
        
        # グラフにデータを追加する
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)

        sheet2.add_chart(chart2, 'A1')

        wb.save(f"{self.dir}売上管理表.xlsx")

        # コンテキスト
        msg = "売上管理表を作成しました"
        context = { 'msg': msg }
        return render(request, self.template, context)
