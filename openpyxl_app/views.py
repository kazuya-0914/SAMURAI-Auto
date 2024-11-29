from django.shortcuts import render

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from datetime import datetime

# Excelファイル保存ディレクトリ
dir = 'openpyxl_app/static/excel/'

# 新規ワークブックを作成
workbook = openpyxl.Workbook()
ws = workbook.active

# --- トップページ --- #
def index(request):
  return render(request, 'openpyxl.html')

# --- 2章 openpyxlの基本的な使い方を学ぼう --- #
# Excel保存
def save(request):
  # 新規Excelファイルの作成
  workbook.save(f"{dir}sample1.xlsx")

  params = {
    'msg_1': 'Excelファイルを保存しました'
  }
  return render(request, 'openpyxl.html', params)

# 既存Excelファイルの読み込みと編集したExcelファイルの保存
def update(request):
  workbook = openpyxl.load_workbook(f"{dir}sample1.xlsx")
  ws = workbook.active

  ws['A1'] = 42
  workbook.save(f"{dir}sample1.xlsx")
  ws['A1'] = 10
  workbook.save(f"{dir}sample2.xlsx")

  params = {
    'msg_1': 'Excelファイルを更新しました'
  }
  return render(request, 'openpyxl.html', params)

# シートの作成
def create(request):
  ws['A1'].value = 'Hello, World!'
  workbook.save(f"{dir}sample3.xlsx")

  params = {
    'msg_1': 'Excelファイルを作成しました'
  }
  return render(request, 'openpyxl.html', params)

# フォーマットの変更
def change(request):
  url_name = request.resolver_match.url_name

  if url_name == 'change-color':
    ws['A1'].fill = PatternFill(patternType='solid', fgColor='FF0000')
    workbook.save(f"{dir}sample3.xlsx")
    text = '色'
  elif url_name == 'change-font':
    font = Font(name='Calibri', size=14, bold=True)
    ws['A1'] = 'Hello, World!'
    ws['A1'].font = font
    workbook.save(f"{dir}sample3.xlsx")
    text = 'フォント'
  else:
    ws['B2'] = 'Hello World!'
    thin_border = Side(style='thin')
    border = Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)
    ws['B2'].border = border
    workbook.save(f"{dir}sample3.xlsx")
    text = '枠線'

  params = {
    'msg_1': f"Excelファイルの{text}を変更しました"
  }
  return render(request, 'openpyxl.html', params)

# --- 3章 openpyxlでデータを整形・加工しよう --- #
# セルの結合
def join(request):
  worksheet = workbook.active
  ws['A1'] = 42
  worksheet.merge_cells('A1:B2')
  workbook.save(f"{dir}sample4.xlsx")

  params = {
    'msg_2': 'Excelファイルのセルを結合しました'
  }
  return render(request, 'openpyxl.html', params)

# セルの結合解除
def cancell(request):
  workbook = openpyxl.load_workbook(f"{dir}sample4.xlsx")
  ws = workbook.active
  ws.unmerge_cells('A1:B2')
  workbook.save(f"{dir}sample4.xlsx")

  params = {
    'msg_2': 'Excelファイルのセル結合を解除しました'
  }
  return render(request, 'openpyxl.html', params)

# 行を挿入
def insert(request):
  ws['B2'] = 68
  workbook.save(f"{dir}sample4.xlsx")
  ws.insert_rows(2, 3)
  workbook.save(f"{dir}sample4.xlsx")

  params = {
    'msg_2': 'Excelファイルに行を挿入しました'
  }
  return render(request, 'openpyxl.html', params)

# 行を削除
def delete_rows(request):
  ws.delete_rows(2, 4)
  workbook.save(f"{dir}sample4.xlsx")

  params = {
    'msg_2': 'Excelファイルの行を削除しました'
  }
  return render(request, 'openpyxl.html', params)

# 新しい行を追加
def append(request):
  ws.append(['Value 1', 'Value 2', 'Value 3'])
  workbook.save(f"{dir}sample4.xlsx")

  params = {
    'msg_2': 'Excelファイルに新しい行を挿入しました'
  }
  return render(request, 'openpyxl.html', params)

# 列を削除
def delete_cols(request):
  ws.delete_cols(2, 2)
  workbook.save(f"{dir}sample4.xlsx")

  params = {
    'msg_2': 'Excelファイルの列を削除しました'
  }
  return render(request, 'openpyxl.html', params)

# ファイルのバックアップ
def backup(request):
  today = datetime.today()
  backup_filename = f"backup_{today.strftime('%Y%m%d')}.xlsx"
  workbook.save(f"{dir}{backup_filename}")

  params = {
    'msg_2': 'Excelファイルのバックアップを行いました'
  }
  return render(request, 'openpyxl.html', params)

# --- 4章 openpyxlでグラフを作成しよう --- #
# データを用意
def prepare(request):
  ws.cell(row=1, column=2, value='d1')
  ws.cell(row=1, column=3, value='d2')

  for i in range(1, 6):
    ws.cell(row=i+1, column=1, value=f"X{i}")
    ws.cell(row=i+1, column=2, value=i * i)
    ws.cell(row=i+1, column=3, value=i % 2)

  workbook.save(f"{dir}output.xlsx")

  params = {
    'msg_3': 'Excelファイルのデータを用意しました'
  }
  return render(request, 'openpyxl.html', params)
  
# グラフを作成
def graph(request):
  url_name = request.resolver_match.url_name
  workbook = openpyxl.load_workbook(f"{dir}output.xlsx")
  ws = workbook.active

  if url_name == 'line':
    chart = LineChart()
    text = '折れ線'
  else:
    chart = BarChart()
    text = '棒'
  chart.title = 'Sample Chart'
  chart.x_axis.title = 'X'
  chart.y_axis.title = 'Y'

  data = Reference(ws, min_row=1, min_col=2, max_row=6, max_col=3)
  chart.add_data(data, titles_from_data=True)
  labels = Reference(ws, min_col=1, min_row=2, max_row=6)
  chart.set_categories(labels)

  chart.legend = None
  ws.add_chart(chart, 'E1')
  workbook.save(f"{dir}output.xlsx")

  params = {
    'msg_3': f"Excelデータの{text}グラフを作成しました"
  }
  return render(request, 'openpyxl.html', params)

# --- 5章 openpyxlで実際にExcelファイルを作成しよう --- #
# 売上表作成
def sales(request):
  header = ['商品名', '単価', '数量', '売上']
  ws.append(header)

  data = [
    ['商品A', 1000, 8, 8000],
    ['商品B', 2000, 5, 10000],
    ['商品C', 500, 13, 6500]
  ]
  for row in data:
    ws.append(row)
  
  thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
  
  for row in ws['A1:D4']:
    for cell in row:
        cell.border = thin_border
  workbook.save(f"{dir}売上表.xlsx")

  params = {
    'msg_4': f"売上表を作成しました"
  }
  return render(request, 'openpyxl.html', params)

# 売上表 棒グラフ
def sales_bar(request):
  workbook = openpyxl.load_workbook(f"{dir}売上表.xlsx")
  ws = workbook.active

  data = Reference(ws, min_row=1, min_col=4, max_row=ws.max_row, max_col=4)
  categories_range = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row, max_col=1)
  chart = BarChart()
  chart.add_data(data, titles_from_data=True)
  chart.set_categories(categories_range)

  chart_cell = ws.max_row + 2
  ws.add_chart(chart, f"A{chart_cell}")
  workbook.save(f"{dir}売上表.xlsx")

  params = {
    'msg_4': f"売上表の棒グラフを作成しました"
  }
  return render(request, 'openpyxl.html', params)